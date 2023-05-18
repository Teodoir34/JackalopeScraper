from tkinter import messagebox
import customtkinter as ctk
from tkinter import filedialog
import praw
import openpyxl
import datetime
# import the required packages and libraries
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from openpyxl.utils.dataframe import dataframe_to_rows
import nltk
from nltk import *
nltk.download('vader_lexicon')
nltk.download('stopwords')
import profanity_check
import nltk







from nltk.sentiment.vader import SentimentIntensityAnalyzer

sid = SentimentIntensityAnalyzer()


#Ted - set the appearance mode to system initially and theme to dark-blue
ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("dark-blue")  # Themes: "blue" (standard), "green", "dark-blue"
#twitSheet = workbook.create_sheet(f"{name}")


#Ted - Create login page class 
class LoginPage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        
        ctk.CTkFrame.__init__(self, parent)
        
        # configure grid layout (4x4)
        self.grid_columnconfigure((0,1,2), weight=1)
        self.grid_rowconfigure((1), weight=0)
        

        # create tabview
        login_entry_frame = ctk.CTkTabview(self, width=250, corner_radius=6)
        login_entry_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        login_entry_frame.add("Login")
        login_entry_frame.add("Register")
        login_entry_frame.tab("Login").grid_columnconfigure(0, weight=1)  # configure grid of individual tabs
        login_entry_frame.tab("Register").grid_columnconfigure(0, weight=1)


        

        username_entry = ctk.CTkEntry(login_entry_frame.tab("Login"), placeholder_text="Username", justify="center")
        username_entry.grid(row=0, column=0, padx=(10, 10), pady=(10, 10), sticky="nsew")

        password_entry = ctk.CTkEntry(login_entry_frame.tab("Login"), placeholder_text="Password",  show='*', justify="center")
        password_entry.grid(row=1, column=0, padx=(10, 10), pady=(10, 10), sticky="nsew")
                    
        def verify():
            """Check if the username and password match an entry in the credentials file."""
            userName = username_entry.get()
            passWord = password_entry.get()
            with open("login.txt", "r") as file:
                for line in file:
                    stored_username, stored_password = line.strip().split(":")
                    if stored_username == userName and stored_password == passWord:
                        controller.show_frame(ScraperPage)
                        return
                else:
                    messagebox.showinfo("Error", "Please provide correct username and password!!")
        
        submit_button = ctk.CTkButton(login_entry_frame.tab("Login"), text="Submit", font=("Arial", 15), command=verify)
        submit_button.grid(row=2, column=0, rowspan=2, padx=10, pady=10, sticky="nsew")
        
        reg_username_entry = ctk.CTkEntry(login_entry_frame.tab("Register"), placeholder_text="Username", justify="center")
        reg_username_entry.grid(row=0, column=0, padx=(10, 10), pady=(10, 10), sticky="nsew")

        reg_password_entry = ctk.CTkEntry(login_entry_frame.tab("Register"), placeholder_text="Password",  show='*', justify="center")
        reg_password_entry.grid(row=1, column=0, padx=(10, 10), pady=(10, 10), sticky="nsew")

        confirm_password_entry = ctk.CTkEntry(login_entry_frame.tab("Register"), placeholder_text="Confirm Password",  show='*', justify="center")
        confirm_password_entry.grid(row=2, column=0, padx=(10, 10), pady=(10, 10), sticky="nsew")
        
        
        def register():
            if reg_username_entry.get()!="" or reg_password_entry.get()!="" or confirm_password_entry.get()!="":
                if reg_password_entry.get()==confirm_password_entry.get():
                    with open("login.txt", "a") as f:
                        f.write(reg_username_entry.get()+":"+reg_password_entry.get()+"\n")
                        messagebox.showinfo("Welcome","Next you'll log in to your twitter and reddit accounts for access")    
                else:
                    messagebox.showinfo("Error","Your password didn't get match!!")
            else:
                messagebox.showinfo("Error", "Please fill the complete field!!")

        def initial_twitter_login():
            #driver = webdriver.Chrome(chrome_options=options)
            driver= webdriver.Chrome(service=Service(ChromeDriverManager().install()))

            #Maximize the Chrome window to full-screen
            driver.maximize_window() 
            loginURL = "https://twitter.com/i/flow/login"
            driver.get(loginURL)
            while(True):
                pass

        twitter_login_button = ctk.CTkButton(login_entry_frame.tab("Register"), text="Login To Twitter", font=("Arial", 15), command=initial_twitter_login)
        twitter_login_button.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")

        submit_button = ctk.CTkButton(login_entry_frame.tab("Register"), text="Register", font=("Arial", 15), command=register)
        submit_button.grid(row=4, column=0, padx=10, pady=10, sticky="nsew")


class ScraperPage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        ctk.CTkFrame.__init__(self, parent)
        
        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=0)
        self.grid_columnconfigure((2, 3), weight=1)
        self.grid_rowconfigure((0), weight=0)
        
        #Ted - Declare variables
        reddit_checkbox_var = ctk.IntVar()
        twitter_checkbox_var = ctk.IntVar()
        chats_checkbox_var = ctk.IntVar()
        cxMed_checkbox_var = ctk.IntVar()
        dxMed_checkbox_var = ctk.IntVar()
        dxSupportTix_checkbox_var = ctk.IntVar()
        twitterAdv_checkbox_var = ctk.IntVar()
        bugscan_checkbox_var = ctk.IntVar()
        tsscan_checkbox_var = ctk.IntVar()
        fileName_entry = ctk.StringVar()    
        workbook = openpyxl.Workbook()

        df = pd.read_excel('dicts.xlsx')
        
        # Select the three columns
        column_1 = df.iloc[:, 0]
        column_2 = df.iloc[:, 1]
        column_3 = df.iloc[:, 2]
        
        # Create three new dataframes
        df1 = pd.DataFrame(column_1)
        df2 = pd.DataFrame(column_2)
        df3 = pd.DataFrame(column_3)

        # Create three new dataframes
        tsFlags = df1.to_dict()
        issueKeys = df2.to_dict()
        hashTags = df3.to_dict()
            
        def clean_text(in_feed):
            # Convert the text to lowercase
            in_feed = in_feed.lower()

            # Remove punctuation
            in_feed = in_feed.replace(',', ' ').replace('.', ' ').replace('!', ' ').replace('?', ' ')

            # Remove stop words
            stopwords = nltk.corpus.stopwords.words('english')
            in_feed = [word for word in in_feed.split() if word not in stopwords]

            # Join the words back together
            in_feed = ' '.join(in_feed)

            return in_feed

        def get_sentiment_score(in_feed):
            # Clean the text
            in_feed = clean_text(in_feed)

            # Get the sentiment score
            sentiment_analyzer = SentimentIntensityAnalyzer()
            sentiment_score = sentiment_analyzer.polarity_scores(in_feed)['compound']

            return sentiment_score
        
        def write_keywords():
            
            keyword_get = self.medKeys_entry.get()
            keywords = keyword_get.split(",")
            ws = workbook.active
            print("ws defined")
            ws.title = "Keywords"

            print("worksheet created")

            column_headings = ["Keywords"]
            ws.append(column_headings)
            name = fileName_entry.get()

            for keyword in keywords:
                ws.append([keyword])
            workbook.save(f'{name}.xlsx')
            print("keyword sheet saved")
        
        def scrape_reddit():
            reddit = praw.Reddit(
                client_id="LA9PBh--URqouEKJVuFDGg",
                client_secret="nauBhC-DEidxjSqId7tOyUYQ8LcRoQ",
                user_agent="Jackalope Scraping")
            
            # Get a subreddit object
            subreddit = reddit.subreddit('SparkDriver')
            name = fileName_entry.get()
            keyword_get = self.medKeys_entry.get()
            keywords = keyword_get.split(", ")

            dict_df = pd.read_excel('dicts.xlsx')
        
            # Select the three columns
            column_1 = dict_df.iloc[:, 0]
            column_2 = dict_df.iloc[:, 1]

            
            # Create three new dataframes
            tsFlags = pd.DataFrame(column_1.to_list())
            tsFlags = tsFlags.dropna()
            tsFlags_string = ", ".join(tsFlags[0].tolist())
            tsFlags = tsFlags_string.split(", ")
            
            issueKeys = pd.DataFrame(column_2.to_list())
            issueKeys = issueKeys.dropna()
            issueKeys_string = ", ".join(issueKeys[0].tolist())
            issueKeys = issueKeys_string.split(", ")


            # Create a worksheet for the post data
            redditSheet = workbook.create_sheet(f'Reddit')
            total = int(self.redditCount_entry.get())
            
            # Get a list of posts from the subreddit
            posts = subreddit.new(limit=total)
                # Recursively iterate through the comments of the post
            def process_comments(comments, parent_comment=None):
                for comment in comments:
                    # Get the desired information about the comment
                    title = post.title
                    bodyComment = comment.body
                    author = comment.author.name if comment.author is not None else 'Deleted'  # Check if the author attribute is None before accessing the name attribute
                    created = datetime.datetime.utcfromtimestamp(comment.created_utc).strftime('%d/%m/%Y %H:%M:%S') # convert created_utc to human-readable format    
                    sentimentComment = get_sentiment_score(bodyComment)
                    if any(keyword in body for keyword in keywords):
                    # Get the list of keywords found in the text
                        keyword_check = [keyword for keyword in keywords if keyword in bodyComment]
                        # Join the keyword list with commas
                        keyword_string = ", ".join(keyword_check)
                    else:
                        keyword_string = "0"
                        
                    # Check if the text contains any keywords from the first dataframe
                    if any(keyword in body for keyword in issueKeys):
                        # Get the list of keywords found in the text
                        issueKeyCheck = [keyword for keyword in issueKeys if keyword in bodyComment]
                        # Join the keyword list with commas
                        issueKey_string = ", ".join(issueKeyCheck)
                    else:
                        issueKey_string = "0"

                    # Check if the text contains any keywords from the second dataframe
                    if any(keyword in body for keyword in tsFlags):
                        # Get the list of keywords found in the text
                        tsFlagCheck = [keyword for keyword in tsFlags if keyword in bodyComment]
                        # Join the keyword list with commas
                        tsFlag_string = ", ".join(tsFlagCheck)
                    else:
                        tsFlag_string = "0"
                    
                    prof_predict = profanity_check.predict_prob([body])


                    # Create a list of values for the comment
                    values = [title, author, bodyComment, created,"www.reddit.com" +  permalink, sentimentComment, keyword_string, issueKey_string, tsFlag_string, str(prof_predict)]

                    # If the comment has a parent comment, add the parent comment's body as an additional value
                    #if parent_comment is not None:
                        #values.append(parent_comment.body)
                    #else:
                    #    values.append("None")

                    # Add the values to the comments worksheet as a new row
                    redditSheet.append(values)

                    # Recursively process the child comments
                    process_comments(comment.replies, comment)
                    workbook.save(f'{name}.xlsx')

                    
# Add column headings to the post worksheet
            column_headings = ['Title', 'Author', 'TXT', 'Created', 'Permalink', 'Sentiment', 'Keywords', 'issueFlag', 'tsFlag', 'Profanity']
            redditSheet.append(column_headings)
            
            
            # Iterate through the list of posts
            for i, post in enumerate(posts):
                # Create a worksheet for the current post and its comments
                title = post.title
                body = post.selftext
                author = post.author.name if post.author is not None else 'Deleted'  # Check if the author attribute is None before accessing the name attribute
                created = datetime.datetime.utcfromtimestamp(post.created_utc).strftime('%d/%m/%Y %H:%M:%S') # convert created_utc to human-readable format
                permalink = post.permalink
                sentiment = get_sentiment_score(body)
                
                # Check if the text contains any keywords
                if any(keyword in body for keyword in keywords):
                    # Get the list of keywords found in the text
                    keyword_check = [keyword for keyword in keywords if keyword in body]
                    # Join the keyword list with commas
                    keyword_string = ", ".join(keyword_check)
                else:
                    keyword_string = "0"
                    
                # Check if the text contains any keywords from the first dataframe
                if any(keyword in body for keyword in issueKeys):
                    # Get the list of keywords found in the text
                    issueKeyCheck = [keyword for keyword in issueKeys if keyword in body]
                    # Join the keyword list with commas
                    issueKey_string = ", ".join(issueKeyCheck)
                else:
                    issueKey_string = "0"

                # Check if the text contains any keywords from the second dataframe
                if any(keyword in body for keyword in tsFlags):
                    # Get the list of keywords found in the text
                    tsFlagCheck = [keyword for keyword in tsFlags if keyword in body]
                    # Join the keyword list with commas
                    tsFlag_string = ", ".join(tsFlagCheck)
                else:
                    tsFlag_string = "0"
                    
                prof_predict = profanity_check.predict_prob([body])

                
                
                
            
                
                # Add the post data to the post worksheet as a new row
                values = [title, author, body, created, "www.reddit.com" + permalink, sentiment, keyword_string, issueKey_string, tsFlag_string, str(prof_predict)]
                redditSheet.append(values)
                process_comments(post.comments, None)

            workbook.save(f'{name}.xlsx')

            print("wb reddit save")
                
        def scrapin_tweets():

            #df = pd.DataFrame()
            name = fileName_entry.get()
            allOfTwit = self.allOf_entry.get()
            exactPhraseTwit = self.exactPhrase_entry.get()
            anyOfTwit = self.anyOf_entry.get()
            sinceTwit = self.since_entry.get()
            untilTwit = self.until_entry.get()
            n = int(self.twitterCount_entry.get())
            hashtagsTwit = "SparkDriver Spark Drive4Walmart Shopping&Delivery S&D SparkShopper SparkShopping"



            options = Options()
            options.add_argument('--no-sandbox')
            #options.add_argument('--headless')
            options.add_argument('user-data-dir=/tmp/tarun')
            options.add_argument("--disable-gpu")
            options.add_argument("--disable-restore-session-state")
            options.add_experimental_option('excludeSwitches', ['enable-logging'])
            #driver = webdriver.Chrome(chrome_options=options)
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

            driver.get("https://twitter.com/search-advanced")
            sleep(10)

            if allOfTwit != '':
                allOfInput = driver.find_element(By.XPATH,"//input[@name='allOfTheseWords']")
                allOfInput.send_keys(f"{allOfTwit}")
            if exactPhraseTwit != '': 
                exactPhraseInput = driver.find_element(By.XPATH,"//input[@name='thisExactPhrase']")
                exactPhraseInput.send_keys(f"{exactPhraseTwit}")
            if anyOfTwit != '':
                anyOfInput = driver.find_element(By.XPATH,"//input[@name='anyOfTheseWords']")
                anyOfInput.send_keys(f"{anyOfTwit}")
                
            hashtagsInput = driver.find_element(By.XPATH,"//input[@name='theseHashtags']")
            hashtagsInput.send_keys(f"{hashtagsTwit}")

            searchButton = driver.find_element(By.XPATH,"//span[contains(text(),'Search')]")
            searchButton.click()

            sleep(5)

            # Get the current URL
            current_url = driver.current_url
            print(current_url)
            current_url = current_url[:-16]
            
            if sinceTwit != '':
                    search_url = current_url + "%20lang%3Aen&src=typed_query&f=live"
            else:
                search_url = current_url + "%20lang%3Aen%20until%3A" + untilTwit + "%20since%3A" + sinceTwit + "&src=typed_query&f=live"
            print(search_url)
            driver.get(f"{search_url}")
            sleep(5)

            UserTags=[]
            Tweets=[]
            TimeStamps=[]
            Sentiments=[]



            articles = driver.find_elements(By.XPATH,"//article[@data-testid='tweet']")

            while len(Tweets) < n:
                for article in articles:
                    try:
                        for i in range(1, n):
                            #add in either way to not grab the adds with no user or way to remove them after the fact - removing before writing best option
                            
                            tweet_handle = driver.find_element(By.XPATH, f'/html/body/div[1]/div/div/div[2]/main/div/div/div/div[1]/div/div[3]/div/section/div/div/div[{i}]/div/div/article/div/div/div[2]/div[2]/div[1]/div/div[1]/div/div/div[2]/div/div[1]/a/div/span').text
                            print("user found")
                            UserTags.append(tweet_handle)
                    except:
                        continue

                    try:

                        TimeStamp = article.find_element(By.XPATH,".//time").get_attribute('datetime')
                        print("time ap")
                        TimeStamps.append(TimeStamp)
                        

                    except:
                        del UserTags[-1]
                        continue

                    try:
                        Tweet = article.find_element(By.XPATH,".//div[@data-testid='tweetText']").text
                        Tweets.append(Tweet)
                        print("twet ap")
                        print(Tweet)
                    except:
                        Tweets.append('')
                        
                    try:
                        in_feed = str(Tweet)
                        in_feed= clean_text(in_feed)
                        Sentiment = get_sentiment_score(in_feed)
                        Sentiments.append(Sentiment)
                    except:
                        Sentiments.append('sentiment didnt work')


                driver.execute_script('window.scrollTo(0,document.body.scrollHeight);')

                sleep(3)

                if len(Tweets) >= n:
                    print("if len >=???")
                    break
                
                articles = driver.find_elements(By.XPATH,"//article[@data-testid='tweet']")
                Tweets = list(set(Tweets))

            df = pd.DataFrame(zip(UserTags,TimeStamps,Tweets, Sentiments),columns=['UserTags','TimeStamps','TXT', 'Sentiment'])
                                            
            print("df zip")



            #convert to openpyxl
            ws = workbook.active
            print("ws defined")
            ws = workbook.create_sheet("Twitter")
            print("worksheet created")
            rows = dataframe_to_rows(df, index=False)
            print("rows to df")

            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                
            workbook.save(f'{name}.xlsx') 
            print("wb twitter save")

            driver.close()

        def cleanChats():

            ws = workbook.active
            name = fileName_entry.get()
            df = pd.read_excel(self.chats_filepath)
            keyword_get = self.medKeys_entry.get()
            print(keyword_get)
            keywords = keyword_get.split(", ")

            dict_df = pd.read_excel('dicts.xlsx')
        
            # Select the three columns
            column_1 = dict_df.iloc[:, 0]
            column_2 = dict_df.iloc[:, 1]
            column_3 = dict_df.iloc[:, 2]
            
            # Create three new dataframes
            tsFlags = pd.DataFrame(column_1.to_list())
            tsFlags = tsFlags.dropna()
            tsFlags_string = ", ".join(tsFlags[0].tolist())
            tsFlags = tsFlags_string.split(", ")
            
            issueKeys = pd.DataFrame(column_2.to_list())
            issueKeys = issueKeys.dropna()
            issueKeys_string = ", ".join(issueKeys[0].tolist())
            issueKeys = issueKeys_string.split(", ")

            hashTags = pd.DataFrame(column_3.to_list())
            hashTags = hashTags.dropna()
            hashTags_string = ", ".join(hashTags[0].tolist())
            hashTags = hashTags_string.split(", ")
            # create string from values in issueKeys
            

            
            
            
            print(keywords)
            print(issueKeys)


            df.columns = [i for i in range(len(df.columns))]
            #fill any blank cells so that the rows don't newline at random points due to empty cells
            df = df[df.columns].fillna('-')

            #drop unneeded columns
            df.drop(columns = [2, 6], axis=1, inplace = True)


            #label columns with easy identifiers
            df.columns = ['date', 'order_num', 'conversation_id', 'timestamp', 'sender', 'TXT']
            print("named columns")
            df['Keywords'] = ""
            df['Sentiment'] = ""
            df['issueFlags'] = ""
            df['tsFlags'] = ""
            df['Profanity'] = ""
            df['# Flags'] = ""


            
            for row in df.iterrows():

                in_feed = row[1]['TXT']
                in_feed = str(in_feed)
                # Analyze the sentiment of the text.
                df.at[row[0], "Sentiment"]  = get_sentiment_score(in_feed)
                print("sentiments")


                keyword_list = [keyword for keyword in keywords if keyword in in_feed]
            # Join the keyword list with commas.
                keyword_string = ", ".join(keyword_list)
                # Write the keyword string to the cell.
                df.at[row[0], "Keywords"] = keyword_string   

                issueKeywords_list = [issue for issue in issueKeys if issue in in_feed]
            # Join the keyword list with commas.
                issueKeyword_string = ", ".join(issueKeywords_list)
                # Write the keyword string to the cell.
                df.at[row[0], "issueFlags"] = issueKeyword_string    
                
                tsKeywords_list = [tsFlag for tsFlag in tsFlags if tsFlag in in_feed]
            # Join the keyword list with commas.
                tsKeywords_string = ", ".join(tsKeywords_list)
                # Write the keyword string to the cell.
                df.at[row[0], "tsFlags"] = tsKeywords_string                     


                predict_prob = profanity_check.predict_prob([in_feed])
                # Write the predict_prob to the 'tsFlags' column
                df.at[row[0], 'Profanity'] = str(predict_prob)

                hashTags_list = [hashTag for hashTag in hashTags if hashTag in in_feed]
            # Join the keyword list with commas.
                hashTags_string = ", ".join(hashTags_list)
                # Write the keyword string to the cell.
                df.at[row[0], "# Flags"] = hashTags_string  

            print("ws defined")
            ws = workbook.create_sheet(f"chats")
            print("worksheet created")
            rows = dataframe_to_rows(df, index=False)
            print("rows to df")

            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    # Check if the issue key column has any text

                    
            workbook.save(f'{name}.xlsx') 
            print("wb chats save")

        def cleanCxMed():

                #read the excel file
            ws = workbook.active
            name = fileName_entry.get()
            df = pd.read_excel(self.cxMed_filepath) 
            keyword_get = self.medKeys_entry.get()
            keywords = keyword_get.split(", ")
            
            dict_df = pd.read_excel('dicts.xlsx')
        
            # Select the three columns
            column_1 = dict_df.iloc[:, 0]
            column_2 = dict_df.iloc[:, 1]
            column_3 = dict_df.iloc[:, 2]
            
            # Create three new dataframes
            tsFlags = pd.DataFrame(column_1.to_list())
            tsFlags = tsFlags.dropna()
            tsFlags_string = ", ".join(tsFlags[0].tolist())
            tsFlags = tsFlags_string.split(", ")
            
            issueKeys = pd.DataFrame(column_2.to_list())
            issueKeys = issueKeys.dropna()
            issueKeys_string = ", ".join(issueKeys[0].tolist())
            issueKeys = issueKeys_string.split(", ")

            hashTags = pd.DataFrame(column_3.to_list())
            hashTags = hashTags.dropna()
            hashTags_string = ", ".join(hashTags[0].tolist())
            hashTags = hashTags_string.split(", ")
            # create string from values in issueKeys

            print("got ts")
            print(keywords)
            
            #get integers for columns to remove columns after next step
            df.columns = [i for i in range(len(df.columns))]
            #fill any blank cells so that the rows don't newline at random points due to empty cells
            df = df[df.columns].fillna('null')

            #drop unneeded columns
            df.drop(columns = [3, 4, 16, 73, 74, 76], axis=1, inplace = True)
            df.drop(df.loc[:, 7:14].columns, axis=1, inplace=True)
            df.drop(df.loc[:, 20:22].columns, axis=1, inplace=True)
            df.drop(df.loc[:, 24:28].columns, axis=1, inplace=True)
            df.drop(df.loc[:, 24:53].columns, axis=1, inplace=True)
            df.drop(df.loc[:, 57:71].columns, axis=1, inplace=True)

            #label columns with easy identifiers
            df.columns = ['survey_id', 'response_date', 'order_date', 'trxn_dt', 'order_num', 'alert_type', 'store_id', 'rec_score', 'overall_score', 'TXT', 'L1_reason', 'L2_reason', 'comb_tops_sent', 'dsp', 'gscope_link', 'CSAT_survey_description']


            print("named columns")
            df['Keywords'] = ""
            df['Sentiment'] = ""
            df['issueFlags'] = ""
            df['tsFlags'] = ""
            df['Profanity'] = ""
            df['# Flags'] = ""

            # Iterate through the rows in the DataFrame.

            for row in df.iterrows():

                in_feed = row[1]['TXT']
                in_feed = str(in_feed)
                
                # Analyze the sentiment of the text.
                df.at[row[0], "Sentiment"]  = get_sentiment_score(in_feed)
                print("sentiments")


                keyword_list = [keyword for keyword in keywords if keyword in in_feed]
            # Join the keyword list with commas.
                keyword_string = ", ".join(keyword_list)
                # Write the keyword string to the cell.
                df.at[row[0], "Keywords"] = keyword_string   

                issueKeywords_list = [issue for issue in issueKeys if issue in in_feed]
            # Join the keyword list with commas.
                issueKeyword_string = ", ".join(issueKeywords_list)
                # Write the keyword string to the cell.
                df.at[row[0], "issueFlags"] = issueKeyword_string    
                
                tsKeywords_list = [tsFlag for tsFlag in tsFlags if tsFlag in in_feed]
            # Join the keyword list with commas.
                tsKeywords_string = ", ".join(tsKeywords_list)
                # Write the keyword string to the cell.
                df.at[row[0], "tsFlags"] = tsKeywords_string                     


                predict_prob = profanity_check.predict_prob([in_feed])
                # Write the predict_prob to the 'tsFlags' column
                df.at[row[0], 'Profanity'] = str(predict_prob)

                hashTags_list = [hashTag for hashTag in hashTags if hashTag in in_feed]
            # Join the keyword list with commas.
                hashTags_string = ", ".join(hashTags_list)
                # Write the keyword string to the cell.
                df.at[row[0], "# Flags"] = hashTags_string  

            print("ws defined")
            ws = workbook.create_sheet(f"CX Medallia")
            print("worksheet created")
            rows = dataframe_to_rows(df, index=False)
            print("rows to df")

            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    # Check if the issue key column has any text

                    
            workbook.save(f'{name}.xlsx')  
            print("wb medallia save")

        def cleanDxMed():

            #read the excel file
            ws = workbook.active
            name = fileName_entry.get()
            df = pd.read_excel(self.dxMed_filepath) 
            keyword_get = self.medKeys_entry.get()
            keywords = keyword_get.split(", ")
            
            dict_df = pd.read_excel('dicts.xlsx')
        
            # Select the three columns
            column_1 = dict_df.iloc[:, 0]
            column_2 = dict_df.iloc[:, 1]
            column_3 = dict_df.iloc[:, 2]
            
            # Create three new dataframes
            tsFlags = pd.DataFrame(column_1.to_list())
            tsFlags = tsFlags.dropna()
            tsFlags_string = ", ".join(tsFlags[0].tolist())
            tsFlags = tsFlags_string.split(", ")
            
            issueKeys = pd.DataFrame(column_2.to_list())
            issueKeys = issueKeys.dropna()
            issueKeys_string = ", ".join(issueKeys[0].tolist())
            issueKeys = issueKeys_string.split(", ")

            hashTags = pd.DataFrame(column_3.to_list())
            hashTags = hashTags.dropna()
            hashTags_string = ", ".join(hashTags[0].tolist())
            hashTags = hashTags_string.split(", ")
            # create string from values in issueKeys

            print("got ts")
            print(keywords)
            
            #get integers for columns to remove columns after next step
            df.columns = [i for i in range(len(df.columns))]
            #fill any blank cells so that the rows don't newline at random points due to empty cells
            df = df[df.columns].fillna('null')

            #drop unneeded columns
            df.drop(columns = [3, 4, 16, 73, 74, 76], axis=1, inplace = True)
            df.drop(df.loc[:, 7:14].columns, axis=1, inplace=True)
            df.drop(df.loc[:, 20:22].columns, axis=1, inplace=True)
            df.drop(df.loc[:, 24:28].columns, axis=1, inplace=True)
            df.drop(df.loc[:, 24:53].columns, axis=1, inplace=True)
            df.drop(df.loc[:, 57:71].columns, axis=1, inplace=True)

            #label columns with easy identifiers
            df.columns = ['survey_id', 'response_date', 'order_date', 'trxn_dt', 'Ord_Num', 'alert_type', 'store_id', 'rec_score', 'overall_score', 'TXT', 'L1_reason', 'L2_reason', 'comb_tops_sent', 'dsp', 'gscope_link', 'CSAT_survey_description']


            print("named columns")
            df['Keywords'] = ""
            df['Sentiment'] = ""
            df['issueFlags'] = ""
            df['tsFLags'] = ""
            df['Profanity'] = ""
            df['# Flags'] = ""

            # Iterate through the rows in the DataFrame.

            for row in df.iterrows():

                # Get the text from the row.
                in_feed = row[1]['TXT']
                in_feed = clean_text(str(in_feed))
                # Analyze the sentiment of the text.
                df.at[row[0], "Sentiment"]  = get_sentiment_score(in_feed)
                print("sentiments")


                keyword_list = [keyword for keyword in keywords if keyword in in_feed]
            # Join the keyword list with commas.
                keyword_string = ", ".join(keyword_list)
                # Write the keyword string to the cell.
                df.at[row[0], "Keywords"] = keyword_string   

                issueKeywords_list = [issue for issue in issueKeys if issue in in_feed]
            # Join the keyword list with commas.
                issueKeyword_string = ", ".join(issueKeywords_list)
                # Write the keyword string to the cell.
                df.at[row[0], "issueFlags"] = issueKeyword_string    
                
                tsKeywords_list = [tsFlag for tsFlag in tsFlags if tsFlag in in_feed]
            # Join the keyword list with commas.
                tsKeywords_string = ", ".join(tsKeywords_list)
                # Write the keyword string to the cell.
                df.at[row[0], "tsFlags"] = tsKeywords_string                     


                predict_prob = profanity_check.predict_prob([in_feed])
                # Write the predict_prob to the 'tsFlags' column
                df.at[row[0], 'Profanity'] = str(predict_prob)

                hashTags_list = [hashTag for hashTag in hashTags if hashTag in in_feed]
            # Join the keyword list with commas.
                hashTags_string = ", ".join(hashTags_list)
                # Write the keyword string to the cell.
                df.at[row[0], "# Flags"] = hashTags_string  

            print("ws defined")
            ws = workbook.create_sheet(f"DX Medallia")
            print("worksheet created")
            rows = dataframe_to_rows(df, index=False)
            print("rows to df")

            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    # Check if the issue key column has any text

                    
            workbook.save(f'{name}.xlsx') 
            print("wb dx medallia save")

        def clean_drvSupportTix():

            #read the excel file
            ws = workbook.active
            name = fileName_entry.get()
            df = pd.read_excel(self.dxSupportTix_filepath) 
            keyword_get = self.medKeys_entry.get()
            keywords = keyword_get.split(", ")
            
            dict_df = pd.read_excel('dicts.xlsx')
        
            # Select the three columns
            column_1 = dict_df.iloc[:, 0]
            column_2 = dict_df.iloc[:, 1]
            column_3 = dict_df.iloc[:, 2]
            
            # Create three new dataframes
            tsFlags = pd.DataFrame(column_1.to_list())
            tsFlags = tsFlags.dropna()
            tsFlags_string = ", ".join(tsFlags[0].tolist())
            tsFlags = tsFlags_string.split(", ")
            
            issueKeys = pd.DataFrame(column_2.to_list())
            issueKeys = issueKeys.dropna()
            issueKeys_string = ", ".join(issueKeys[0].tolist())
            issueKeys = issueKeys_string.split(", ")

            hashTags = pd.DataFrame(column_3.to_list())
            hashTags = hashTags.dropna()
            hashTags_string = ", ".join(hashTags[0].tolist())
            hashTags = hashTags_string.split(", ")
            # create string from values in issueKeys

            print("got ts")
            print(keywords)
            
            #get integers for columns to remove columns after next step
            df.columns = [i for i in range(len(df.columns))]
            #fill any blank cells so that the rows don't newline at random points due to empty cells
            df = df[df.columns].fillna('null')

            #drop unneeded columns
            df.drop(columns = [0, 5], axis=1, inplace = True)

            #label columns with easy identifiers
            df.columns = ['Ref Num', 'Ord_Num', 'L1_reason', 'L1_reason', 'storeID', 'agentID', 'dxID', 'dxPhone', 'Queue Name', 'TXT']


            print("named columns")
            df['Keywords'] = ""
            df['Sentiment'] = ""
            df['issueFlags'] = ""
            df['Profanity'] = ""
            df['tsFLags'] = ""
            df['# Flags'] = ""


            # Iterate through the rows in the DataFrame.

            for row in df.iterrows():


                # Get the text from the row.
                in_feed = row[1]['TXT']
                in_feed = clean_text(str(in_feed))
                # Analyze the sentiment of the text.
                df.at[row[0], "Sentiment"]  = get_sentiment_score(in_feed)
                print("sentiments")


                keyword_list = [keyword for keyword in keywords if keyword in in_feed]
            # Join the keyword list with commas.
                keyword_string = ", ".join(keyword_list)
                # Write the keyword string to the cell.
                df.at[row[0], "Keywords"] = keyword_string   

                issueKeywords_list = [issue for issue in issueKeys if issue in in_feed]
            # Join the keyword list with commas.
                issueKeyword_string = ", ".join(issueKeywords_list)
                # Write the keyword string to the cell.
                df.at[row[0], "issueFlags"] = issueKeyword_string    
                
                tsKeywords_list = [tsFlag for tsFlag in tsFlags if tsFlag in in_feed]
            # Join the keyword list with commas.
                tsKeywords_string = ", ".join(tsKeywords_list)
                # Write the keyword string to the cell.
                df.at[row[0], "tsFlags"] = tsKeywords_string                     


                predict_prob = profanity_check.predict_prob([in_feed])
                # Write the predict_prob to the 'tsFlags' column
                df.at[row[0], 'Profanity'] = str(predict_prob)

                hashTags_list = [hashTag for hashTag in hashTags if hashTag in in_feed]
            # Join the keyword list with commas.
                hashTags_string = ", ".join(hashTags_list)
                # Write the keyword string to the cell.
                df.at[row[0], "# Flags"] = hashTags_string  

            print("ws defined")
            ws = workbook.create_sheet(f"DX Support")
            print("worksheet created")
            rows = dataframe_to_rows(df, index=False)
            print("rows to df")

            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    # Check if the issue key column has any text

                    
            workbook.save(f'{name}.xlsx') 
            print("wb dx medallia save")
            print("wb dx support tix save")

        def scrape_all():

            redditAsk = self.reddit_checkbox.get()
            twitterAsk = self.twitter_checkbox.get()
            chatsAsk = self.chats_checkbox.get()
            cxMedAsk = self.cxMed_checkbox.get()
            dxMedAsk = self.dxMed_checkbox.get()
            dxSupportAsk = self.dxSupportTix_checkbox.get()
            medKeys = self.medKeys_entry.get()

            bugScanAsk = self.bugScan_checkbox.get()
            tsScanAsk = self.tsScan_checkbox.get()

            name = fileName_entry.get()
            
            

            
            if medKeys != "":
                print("keywords")
                write_keywords()
                print("write keys")
                

            if redditAsk == "on":
                print("now reddit")
                scrape_reddit()
                print("reddit")
            

            if twitterAsk == "on":
                print("now tweets")
                scrapin_tweets()
                print("tweets")
                
            if chatsAsk == "on":
                print("now chats")
                cleanChats()
                print("chats")
                

            if cxMedAsk == "on":
                print("now medallia")
                cleanCxMed()
                print("medallia")

            if dxMedAsk == "on":
                print("now driver medallia")
                cleanDxMed()
                print("driver medallia")

            if dxSupportAsk == "on":
                print("now driver medallia")
                clean_drvSupportTix()
                print("driver support")

            if bugScanAsk == "on":
                print("bug scan")


            if tsScanAsk == "on":
                print("ts scan")
                
                


            workbook.save(f'{name}.xlsx')

            print("scraping done!")
            
        self.title_frame = ctk.CTkFrame(self, fg_color="orange", corner_radius=10)
        self.title_frame.grid(row=0, column=0, columnspan=4, padx=10, pady=10,  sticky="nsew")
        self.title_frame.grid_columnconfigure((0), weight=1)
        self.title_frame.grid_rowconfigure((0), weight=1)
        logo_label = ctk.CTkLabel(self.title_frame, text="Jackalope Social & Feedback Analysis Tool", font=ctk.CTkFont(size=20, weight="bold"))
        logo_label.grid(row=0, column=0, padx=10, pady=10)
        logo_label.configure(width=1000)
        

        #Ted - create sidebar frame with widgets
        self.sidebar_frame = ctk.CTkFrame(self, fg_color="pink", width=800, height=900, corner_radius=10)
        self.sidebar_frame.configure(width=400)
        self.sidebar_frame.grid(row=1, column=0, rowspan=4, padx=10, pady=10,  sticky="nsw")
        self.sidebar_frame.grid_columnconfigure((0), weight=1)
        self.sidebar_frame.grid_rowconfigure((0), weight=1)

        
        #Ted - create checkbox options frame and widget    
        self.cb_frame = ctk.CTkFrame(self.sidebar_frame, corner_radius=6)
        self.cb_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        self.cb_frame.grid_columnconfigure((0), weight=1)
        self.cb_frame_label = ctk.CTkLabel(self.cb_frame, text="What will you be scraping?:")
        self.cb_frame_label.grid(row=0, column=0, padx=20, pady=(10, 10), sticky="nsew")
        self.reddit_checkbox = ctk.CTkCheckBox(self.cb_frame, text="Reddit", command=self.reddit_checkbox_var, variable=reddit_checkbox_var,  onvalue="on", offvalue="off")
        self.reddit_checkbox.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        self.twitter_checkbox = ctk.CTkCheckBox(self.cb_frame, text="Twitter", command=self.twitter_checkbox_var, variable=twitter_checkbox_var,  onvalue="on", offvalue="off")
        self.twitter_checkbox.grid(row=2, column=0, padx=20, pady=10,sticky="nsew")
        self.chats_checkbox = ctk.CTkCheckBox(self.cb_frame, text="Chats", command=self.chats_checkbox_var, variable=chats_checkbox_var,  onvalue="on", offvalue="off")
        self.chats_checkbox.grid(row=3, column=0, padx=20, pady=10, sticky="nsew")
        self.cxMed_checkbox = ctk.CTkCheckBox(self.cb_frame, text="Medallia", command=self.cxMed_checkbox_var, variable=cxMed_checkbox_var,  onvalue="on", offvalue="off")
        self.cxMed_checkbox.grid(row=4, column=0, padx=20, pady=10, sticky="nsew")
        self.dxMed_checkbox = ctk.CTkCheckBox(self.cb_frame, text="Driver Medallia", command=self.dxMed_checkbox_var, variable=dxMed_checkbox_var,  onvalue="on", offvalue="off")
        self.dxMed_checkbox.grid(row=5, column=0, padx=20, pady=10, sticky="nsew")
        self.dxSupportTix_checkbox = ctk.CTkCheckBox(self.cb_frame, text="Driver Support", command=self.dxSupportTix_checkbox_var, variable=dxSupportTix_checkbox_var,  onvalue="on", offvalue="off")
        self.dxSupportTix_checkbox.grid(row=6, column=0, padx=20, pady=10, sticky="nsew")
        

        self.optionsFrame = ctk.CTkFrame(self.sidebar_frame, corner_radius=6)
        self.optionsFrame.grid(row=1, column=0,  padx=10, pady=10, sticky="nsew")
        self.optionsFrame.grid_columnconfigure((0), weight=1)
        self.optionsFrame_label = ctk.CTkLabel(self.optionsFrame, text="Advanced Options:")
        self.optionsFrame_label.grid(row=0, column=0,  padx=10, pady=0, sticky="nsew")
        self.twitterAdv_checkbox = ctk.CTkCheckBox(self.optionsFrame, text="Twitter Advanced", command=self.twitterAdv_checkbox_var, variable=twitterAdv_checkbox_var,  onvalue="on", offvalue="off")
        self.twitterAdv_checkbox.grid(row=1, column=0, padx=20, pady=10,sticky="nsew")
        self.bugScan_checkbox = ctk.CTkCheckBox(self.optionsFrame, text="Bug Scan", command=self.bugscan_checkbox_var, variable=bugscan_checkbox_var,  onvalue="on", offvalue="off")
        self.bugScan_checkbox.grid(row=2, column=0, padx=20, pady=10, sticky="nsew")
        self.tsScan_checkbox = ctk.CTkCheckBox(self.optionsFrame, text="T&S Scan", command=self.tsscan_checkbox_var, variable=tsscan_checkbox_var,  onvalue="on", offvalue="off")
        self.tsScan_checkbox.grid(row=3, column=0, padx=20, pady=10, sticky="nsew")
        self.tsScan_checkbox.configure(width=100)
        
        #Ted - create appearance adjustment frame and widget
        self.changes_frame = ctk.CTkFrame(self.sidebar_frame, corner_radius=6)
        self.changes_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
        self.changes_frame.grid_columnconfigure((0), weight=1)
        self.appearance_mode_label = ctk.CTkLabel(self.changes_frame, text="Appearance Mode:")
        self.appearance_mode_label.grid(row=1, column=0, padx=20, pady=(10, 10), sticky="nsew")
        self.appearance_mode_optionemenu = ctk.CTkOptionMenu(self.changes_frame, values=["Light", "Dark", "System"], command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=2, column=0, padx=20, pady=(5, 15), sticky="nsew")
        self.scaling_label = ctk.CTkLabel(self.changes_frame, text="Scaling:")
        self.scaling_label.grid(row=3, column=0, padx=20, pady=(15, 5), sticky="nsew")
        self.scaling_optionemenu = ctk.CTkOptionMenu(self.changes_frame, values=["80%", "90%", "100%", "110%", "120%"], command=self.change_scaling_event)
        self.scaling_optionemenu.grid(row=4, column=0, padx=20, pady=(10, 10), sticky="sew")

        #Ted - create sidebar frame with widgets
        self.input_frame = ctk.CTkFrame(self, corner_radius=10)
        self.input_frame.grid(row=1, column=1, rowspan=4, padx=10, pady=10,  sticky="nsew")
        self.input_frame.configure(fg_color="purple")

        self.excelFileScraper = ctk.CTkFrame(self.input_frame,  corner_radius=10)
        self.excelFileScraper.grid(row=0, column=1, columnspan = 2, padx=10, pady=(15, 5), sticky="nsew")
        self.excelFileScraper.grid_rowconfigure(0, weight=1)  # configure grid system
        self.excelFileScraper.grid_columnconfigure((0, 1), weight=1)
        self.excelFileScraper_label = ctk.CTkLabel(self.excelFileScraper, text="Excel File Scraper")
        self.excelFileScraper_label.grid(row=0, column=0, columnspan = 2, padx=10, pady=0,  sticky="nsew")
        self.medKeys_entry = ctk.CTkEntry(self.excelFileScraper, placeholder_text="Comma Separated Keywords. Ex: Spark, bugs, issue")
        self.medKeys_entry.grid(row=1, column=0, columnspan = 2,padx=10, pady=(5, 15), sticky="nsew")
        self.upload_LiveChats = ctk.CTkButton(self.excelFileScraper, text="Upload LIVE chat .xlsx", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.browseChatFiles)
        self.upload_LiveChats.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
        self.upload_DxSupportTix = ctk.CTkButton(self.excelFileScraper, text="Upload DX Support Tix .xlsx", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.browseDxSupportTixFiles)
        self.upload_DxSupportTix.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
        self.upload_CxMed = ctk.CTkButton(self.excelFileScraper, text="Upload CX Medallia .xlsx", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.browseCxMedFiles)
        self.upload_CxMed.grid(row=2, column=1, padx=10, pady=10, sticky="nsew")
        self.upload_DxMed = ctk.CTkButton(self.excelFileScraper, text="Upload DX Med .xlsx", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.browseDxMedFiles)
        self.upload_DxMed.grid(row=3, column=1, padx=10, pady=10, sticky="nsew")

        #Ted - create input frame and widgets
        self.redditFrame = ctk.CTkFrame(self.input_frame, corner_radius=10)
        self.redditFrame.grid(row=1, column=1, padx=10, pady=(15, 5), sticky="nsew")
        self.redditFrame.grid_rowconfigure(0, weight=1)  # configure grid system
        self.redditFrame.grid_columnconfigure(0, weight=1)
        self.redditFrame_label = ctk.CTkLabel(self.redditFrame, text="Reddit")
        self.redditFrame_label.grid(row=0, column=0,  padx=10, pady=0, sticky="nsew")
        self.redditCount_entry = ctk.CTkEntry(self.redditFrame, placeholder_text="# of posts from /r/SparkDriver/new/")
        self.redditCount_entry.grid(row=1, column=0, padx=10, pady=(5, 15), sticky="nsew")
        self.redditCount_entry.configure(width=800)

        self.twitterFrame = ctk.CTkFrame(self.input_frame, corner_radius=10)
        self.twitterFrame.grid(row=2, column=1, padx=10, pady=(15, 5), sticky="nsew")
        self.twitterFrame.grid_rowconfigure(0, weight=1)  # configure grid system
        self.twitterFrame.grid_columnconfigure((0,1), weight=1)
        self.twitterFrame_label = ctk.CTkLabel(self.twitterFrame, text="Twitter")
        self.twitterFrame_label.grid(row=0, column=0,columnspan=2,  padx=10, pady=0, sticky="nsew")
        self.anyOf_entry = ctk.CTkEntry(self.twitterFrame, placeholder_text="Any of these words: please enter separated by spaces")
        self.anyOf_entry.grid(row=1, column=0, columnspan=2, padx=10, pady=(5, 15), sticky="nsew")
        self.allOf_entry = ctk.CTkEntry(self.twitterFrame, placeholder_text="All of these words: please enter separated by spaces")
        self.allOf_entry.grid(row=2, column=0,columnspan=2, padx=10, pady=(5, 15), sticky="nsew")
        self.exactPhrase_entry = ctk.CTkEntry(self.twitterFrame, placeholder_text="This exact phrase: only one phrase at a time")
        self.exactPhrase_entry.grid(row=3, column=0,columnspan=2, padx=10, pady=(5, 15), sticky="nsew")
        self.since_entry = ctk.CTkEntry(self.twitterFrame, placeholder_text="Since: YYYY-MM-DD")
        self.since_entry.grid(row=4, column=0, padx=10, pady=(5, 15), sticky="nsew")
        self.until_entry = ctk.CTkEntry(self.twitterFrame, placeholder_text="Until: YYYY-MM-DD")
        self.until_entry.grid(row=4, column=1, padx=10, pady=(5, 15), sticky="nsew")
        
        self.twitterCount_entry = ctk.CTkEntry(self.twitterFrame, placeholder_text="# of tweets")
        self.twitterCount_entry.grid(row=5, column=0,columnspan=2, padx=10, pady=(5, 15), sticky="nsew")


        #Ted - create more input frame widgets
        self.nameScrapeFrame = ctk.CTkFrame(self.input_frame,corner_radius=10)
        self.nameScrapeFrame.grid(row=3, column=1, padx=10, pady=(15, 5), sticky="nsew")
        self.nameScrapeFrame.grid_rowconfigure(0, weight=1)  # configure grid system
        self.nameScrapeFrame.grid_columnconfigure(0, weight=1)
        nameScrapeFrame_label = ctk.CTkLabel(self.nameScrapeFrame , text="Name and Scrape!")
        nameScrapeFrame_label.grid(row=0, column=0, padx=10, pady=0,  sticky="nsew")
        fileName_entry = ctk.CTkEntry(self.nameScrapeFrame , placeholder_text="Enter File Name: All scrapes will save to the same excel workbook")
        fileName_entry.grid(row=1, column=0, padx=10, pady=(5, 15), sticky="nsew")
        scrape_button = ctk.CTkButton(self.nameScrapeFrame , text="Scrape!",  border_width=2, text_color=("gray10", "#DCE4EE"), command=scrape_all)
        scrape_button.grid(row=2, column=0, padx=10, pady=(5, 15), sticky="nsew")
        twitterLink_button = ctk.CTkButton(self.nameScrapeFrame , text="Twitter Link for when API Cap Reached",  border_width=2, text_color=("gray10", "#DCE4EE"), command=self.open_twitterlink)
        twitterLink_button.grid(row=3, column=0, padx=10, pady=(5, 15), sticky="nsew")

        
        
        #Ted - set initial options
        self.redditCount_entry.configure(state="disabled")
        self.allOf_entry.configure(state="disabled")
        self.exactPhrase_entry.configure(state="disabled")
        self.anyOf_entry.configure(state="disabled")
        self.twitterCount_entry.configure(state="disabled")
        self.upload_LiveChats.configure(state="disabled")
        self.upload_DxSupportTix.configure(state="disabled")
        self.upload_CxMed.configure(state="disabled")
        self.upload_DxMed.configure(state="disabled")
        self.appearance_mode_optionemenu.set("Dark")
        self.scaling_optionemenu.set("100%")
    
    def browseChatFiles(self):
        # Create a new window for file dialog box
        file_window = ctk.CTkToplevel(app)
        file_window.withdraw()
        # Open file dialog box
        chats_filepath = filedialog.askopenfilename(parent=file_window)
        self.chats_filepath = chats_filepath
        self.upload_LiveChats.configure(text="Successful")

    def browseCxMedFiles(self):
        # Create a new window for file dialog box
        file_window = ctk.CTkToplevel(app)
        file_window.withdraw()
        # Open file dialog box
        cxMed_filepath = filedialog.askopenfilename(parent=file_window)
        self.cxMed_filepath = cxMed_filepath
        self.upload_CxMed.configure(text="Successful")

    def browseDxMedFiles(self):
        # Create a new window for file dialog box
        file_window = ctk.CTkToplevel(app)
        file_window.withdraw()
        # Open file dialog box
        dxMed_filepath = filedialog.askopenfilename(parent=file_window)
        self.dxMed_filepath = dxMed_filepath
        self.upload_DxMed.configure(text="Successful")

    def browseDxSupportTixFiles(self):
        # Create a new window for file dialog box
        file_window = ctk.CTkToplevel(app)
        file_window.withdraw()
        # Open file dialog box
        dxSupportTix_filepath = filedialog.askopenfilename(parent=file_window)
        self.dxSupportTix_filepath = dxSupportTix_filepath
        self.upload_DxSupportTix.configure(text="Successful")

    def open_input_dialog_event(self):
        dialog = ctk.CTkInputDialog(text="Type in a number:", title="CTkInputDialog")
        print("CTkInputDialog:", dialog.get_input())
    
    def change_appearance_mode_event(self, new_appearance_mode: str):
        ctk.set_appearance_mode(new_appearance_mode)
    
    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        ctk.set_widget_scaling(new_scaling_float)

    def sidebar_checkbox_event(self):
        print("sidebar_button click")
        
    def reddit_checkbox_var(self):
        self.redditCount_entry.configure(state="normal")
        print("sidebar_button click")
            
    def twitter_checkbox_var(self):
        self.anyOf_entry.configure(state="normal")
        self.twitterCount_entry.configure(state="normal")
        print("sidebar_button click")
        
    def twitterAdv_checkbox_var(self):
        messagebox.showinfo("Warning", "Esing advanced search will provide more specific but limited results!")
        self.allOf_entry.configure(state="normal")
        self.exactPhrase_entry.configure(state="normal")
        self.anyOf_entry.configure(state="normal")
        self.until_entry.configure(state="normal")
        self.since_entry.configure(state="normal")
        self.twitterCount_entry.configure(state="normal")
        print("sidebar_button click")
        
    def chats_checkbox_var(self):
        self.medKeys_entry.configure(state="normal")
        self.upload_LiveChats.configure(state="enabled")
        print("sidebar_button click")
        
    def cxMed_checkbox_var(self):
        self.medKeys_entry.configure(state="normal")
        self.upload_CxMed.configure(state="enabled")
        print("sidebar_button click")

    def dxMed_checkbox_var(self):
        self.medKeys_entry.configure(state="normal")
        self.upload_DxMed.configure(state="enabled")
        print("sidebar_button click")

    def dxSupportTix_checkbox_var(self):
        self.medKeys_entry.configure(state="normal")
        self.upload_DxSupportTix.configure(state="enabled")
        print("sidebar_button click")

    def bugscan_checkbox_var(self):
        bugs = "on"
        print("sidebar_button click")
    
    def tsscan_checkbox_var(self):
        tsscan = "on"
        print("sidebar_button click")
        
    def open_twitterlink(event):
        print("twitter link button pressed")
        open("https://www.google.com")

    
class Application(ctk.CTk):
    def __init__(self, *args, **kwargs):
        ctk.CTk.__init__(self, *args, **kwargs)
        
        self.geometry("{}x{}".format(int(self.winfo_screenwidth()), int(self.winfo_screenheight())))
        window = ctk.CTkFrame(self,  corner_radius=10)
        window.pack(side="top", expand=True)
        window.configure(width=1200, height=1000)

        self.frames = {}
        for F in (LoginPage, ScraperPage):
            frame = F(window, self)
            self.frames[F] = frame
            frame.grid(row = 0, column=0, sticky="nsew")
            #frame.configure(width=1200, height=1000)
            frame.configure(fg_color="transparent")

        self.show_frame(ScraperPage)
        
    def show_frame(self, page):
        frame = self.frames[page]
        frame.tkraise()
        self.title("Jackalope Keyword & Sentiment Analysis Tool")

app = Application()
app.maxsize(1600,1100)
app.mainloop()

